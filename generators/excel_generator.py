"""
Excel Workbook Generator for Equity Research Reports.

This module generates comprehensive Excel workbooks (.xlsx) with
multiple sheets containing all financial data, calculations, and analysis.
"""

import sys
from pathlib import Path
from datetime import datetime
from typing import Dict, Any
import pandas as pd

# Add project root to path
sys.path.insert(0, str(Path(__file__).parent.parent))

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from agents.state import EquityResearchState
from utils.logger import logger


def generate_excel_workbook(state: EquityResearchState, output_dir: str = "output") -> str:
    """
    Generate comprehensive Excel workbook with all data and analysis.
    
    Creates a multi-sheet .xlsx file with:
    1. Summary - Key metrics and recommendation
    2. Financial Ratios - All calculated ratios
    3. Income Statement - Historical data
    4. Balance Sheet - Historical data
    5. Cash Flow - Historical data
    6. Stock Prices - Price history
    7. Dividends - Dividend history
    8. Valuation - Beta, CAPM, DDM calculations
    9. News - Recent developments
    
    Args:
        state: Complete EquityResearchState with all data and analysis
        output_dir: Directory to save the workbook (default: "output")
    
    Returns:
        str: Path to generated Excel workbook
    
    Example:
        >>> from agents import run_research_workflow
        >>> state = run_research_workflow("RELIANCE")
        >>> path = generate_excel_workbook(state)
        >>> print(f"Workbook saved: {path}")
    """
    logger.info(f"\n{'='*70}")
    logger.info(f"📊 GENERATING EXCEL WORKBOOK: {state['company_name']} ({state['ticker']})")
    logger.info(f"{'='*70}\n")
    
    # Create output directory
    output_path = Path(output_dir)
    output_path.mkdir(parents=True, exist_ok=True)
    
    # Initialize workbook
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet
    
    # 1. Summary Sheet
    logger.info("📝 Step 1/9: Creating Summary sheet...")
    _add_summary_sheet(wb, state)
    
    # 2. Financial Ratios
    logger.info("📝 Step 2/9: Adding Financial Ratios sheet...")
    _add_ratios_sheet(wb, state)
    
    # 3. Income Statement
    logger.info("📝 Step 3/9: Adding Income Statement sheet...")
    _add_income_statement_sheet(wb, state)
    
    # 4. Balance Sheet
    logger.info("📝 Step 4/9: Adding Balance Sheet sheet...")
    _add_balance_sheet_sheet(wb, state)
    
    # 5. Cash Flow
    logger.info("📝 Step 5/9: Adding Cash Flow sheet...")
    _add_cash_flow_sheet(wb, state)
    
    # 6. Stock Prices
    logger.info("📝 Step 6/9: Adding Stock Prices sheet...")
    _add_stock_prices_sheet(wb, state)
    
    # 7. Dividends
    logger.info("📝 Step 7/9: Adding Dividends sheet...")
    _add_dividends_sheet(wb, state)
    
    # 8. Valuation
    logger.info("📝 Step 8/9: Adding Valuation sheet...")
    _add_valuation_sheet(wb, state)
    
    # 9. News
    logger.info("📝 Step 9/9: Adding News sheet...")
    _add_news_sheet(wb, state)
    
    # Save workbook
    filename = f"Equity_Research_Data_{state['ticker']}_{datetime.now().strftime('%Y%m%d')}.xlsx"
    filepath = output_path / filename
    wb.save(str(filepath))
    
    logger.success(f"✅ Excel workbook generated: {filepath}")
    logger.info(f"   File size: {filepath.stat().st_size / 1024:.2f} KB")
    logger.info(f"   Sheets: {len(wb.sheetnames)}")
    
    return str(filepath)


def _style_header_row(ws, row_num: int, num_cols: int):
    """Apply styling to header row."""
    fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    font = Font(bold=True, color="FFFFFF")
    alignment = Alignment(horizontal="center", vertical="center")
    
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.fill = fill
        cell.font = font
        cell.alignment = alignment


def _add_summary_sheet(wb: Workbook, state: EquityResearchState):
    """Add summary sheet with key metrics."""
    ws = wb.create_sheet("Summary", 0)
    
    # Title
    ws['A1'] = "EQUITY RESEARCH SUMMARY"
    ws['A1'].font = Font(bold=True, size=16, color="366092")
    ws.merge_cells('A1:D1')
    
    # Company info
    company_info = state.get('company_info', {})
    stock_prices = state.get('stock_prices')
    
    ws['A3'] = "Company:"
    ws['B3'] = state.get('company_name', state['ticker'])
    ws['A4'] = "Ticker:"
    ws['B4'] = state['ticker']
    ws['A5'] = "Sector:"
    ws['B5'] = company_info.get('sector', 'N/A')
    ws['A6'] = "Industry:"
    ws['B6'] = company_info.get('industry', 'N/A')
    ws['A7'] = "Report Date:"
    ws['B7'] = datetime.now().strftime("%Y-%m-%d")
    
    # Style labels
    for row in range(3, 8):
        ws.cell(row=row, column=1).font = Font(bold=True)
    
    # Market data
    ws['A9'] = "MARKET DATA"
    ws['A9'].font = Font(bold=True, size=14, color="366092")
    
    current_price = stock_prices['Close'].iloc[-1] if stock_prices is not None and not stock_prices.empty else 0
    market_cap = company_info.get('marketCap', 0) / 1e9 if company_info.get('marketCap') else 0
    
    ws['A10'] = "Current Price:"
    ws['B10'] = f"₹{current_price:.2f}"
    ws['A11'] = "Market Cap:"
    ws['B11'] = f"₹{market_cap:.2f}B"
    ws['A12'] = "52-Week High:"
    ws['B12'] = f"₹{company_info.get('fiftyTwoWeekHigh', 'N/A')}"
    ws['A13'] = "52-Week Low:"
    ws['B13'] = f"₹{company_info.get('fiftyTwoWeekLow', 'N/A')}"
    
    for row in range(10, 14):
        ws.cell(row=row, column=1).font = Font(bold=True)
    
    # Valuation metrics
    ws['A15'] = "VALUATION METRICS"
    ws['A15'].font = Font(bold=True, size=14, color="366092")
    
    beta = state.get('beta', 0)
    cost_of_equity = state.get('cost_of_equity', 0)
    ddm = state.get('ddm_valuation', {})
    
    ws['A16'] = "Beta:"
    ws['B16'] = f"{beta:.3f}" if beta else "N/A"
    ws['A17'] = "Cost of Equity:"
    ws['B17'] = f"{cost_of_equity:.2%}" if cost_of_equity else "N/A"
    ws['A18'] = "Fair Value (DDM):"
    ws['B18'] = f"₹{ddm.get('fair_value', 0):.2f}" if ddm and ddm.get('applicable') else "N/A"
    ws['A19'] = "Upside/Downside:"
    ws['B19'] = f"{ddm.get('upside_downside', 0):.1%}" if ddm and ddm.get('applicable') else "N/A"
    
    for row in range(16, 20):
        ws.cell(row=row, column=1).font = Font(bold=True)
    
    # Recommendation
    ws['A21'] = "RECOMMENDATION"
    ws['A21'].font = Font(bold=True, size=14, color="366092")
    
    recommendation = state.get('valuation_recommendation', 'N/A')
    ws['A22'] = recommendation
    ws['A22'].font = Font(bold=True, size=12)
    
    # Color code recommendation
    if 'Buy' in recommendation:
        ws['A22'].font = Font(bold=True, size=12, color="008000")  # Green
    elif 'Sell' in recommendation:
        ws['A22'].font = Font(bold=True, size=12, color="FF0000")  # Red
    else:
        ws['A22'].font = Font(bold=True, size=12, color="FFA500")  # Orange
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 30


def _add_ratios_sheet(wb: Workbook, state: EquityResearchState):
    """Add financial ratios sheet with year-on-year comparison."""
    ws = wb.create_sheet("Financial Ratios")
    
    ratios_by_year = state.get('ratios_by_year', [])
    ratios = state.get('ratios', {})  # Fallback to latest period
    
    if not ratios_by_year and not ratios:
        ws['A1'] = "No ratio data available"
        return
    
    # Title
    ws['A1'] = "FINANCIAL RATIOS ANALYSIS (Year-on-Year)"
    ws['A1'].font = Font(bold=True, size=14, color="366092")
    
    # If we have year-on-year data, show multi-period comparison
    if ratios_by_year:
        # Merge title across all columns
        num_periods = len(ratios_by_year)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_periods + 1)
        
        # Add column headers (dates)
        ws.cell(row=3, column=1, value="Ratio").font = Font(bold=True, color="FFFFFF")
        ws.cell(row=3, column=1).fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        for idx, year_data in enumerate(ratios_by_year):
            col = idx + 2
            # Extract year from date (e.g., "2024-03-31" -> "FY 2024")
            date_str = year_data.get('date', '')
            year = date_str[:4] if date_str else f"Period {idx}"
            ws.cell(row=3, column=col, value=f"FY {year}").font = Font(bold=True, color="FFFFFF")
            ws.cell(row=3, column=col).fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            ws.cell(row=3, column=col).alignment = Alignment(horizontal="center")
        
        row = 4
        
        # Define ratio categories
        categories = [
            ('LIQUIDITY RATIOS', ['current_ratio', 'quick_ratio', 'cash_ratio']),
            ('EFFICIENCY RATIOS', ['asset_turnover', 'inventory_turnover', 'receivables_turnover', 'days_sales_outstanding']),
            ('SOLVENCY/LEVERAGE RATIOS', ['debt_to_equity', 'debt_ratio', 'interest_coverage', 'equity_multiplier']),
            ('PROFITABILITY RATIOS', ['gross_profit_margin', 'operating_profit_margin', 'net_profit_margin', 'return_on_assets', 'return_on_equity', 'return_on_invested_capital']),
            ('VALUATION RATIOS', ['pe_ratio', 'pb_ratio', 'dividend_yield'])
        ]
        
        for category_name, ratio_names in categories:
            # Category header
            ws.cell(row=row, column=1, value=category_name).font = Font(bold=True, size=11, color="366092")
            row += 1
            
            # Add ratios for this category
            for ratio_name in ratio_names:
                ws.cell(row=row, column=1, value=ratio_name.replace('_', ' ').title())
                
                # Add values for each period
                for idx, year_data in enumerate(ratios_by_year):
                    col = idx + 2
                    ratio_value = year_data.get('ratios', {}).get(ratio_name)
                    
                    if ratio_value is not None:
                        # Format percentage ratios
                        if 'margin' in ratio_name or 'return' in ratio_name:
                            ws.cell(row=row, column=col, value=f"{ratio_value:.2f}%")
                        else:
                            ws.cell(row=row, column=col, value=f"{ratio_value:.2f}")
                        ws.cell(row=row, column=col).alignment = Alignment(horizontal="right")
                    else:
                        ws.cell(row=row, column=col, value="N/A")
                        ws.cell(row=row, column=col).alignment = Alignment(horizontal="center")
                
                row += 1
            
            row += 1  # Blank row between categories
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 35
        for idx in range(num_periods):
            col_letter = chr(66 + idx)  # B, C, D, etc.
            ws.column_dimensions[col_letter].width = 15
    
    else:
        # Fallback to single-period display (old format)
        ws.merge_cells('A1:C1')
        row = 3
        
        # Liquidity Ratios
        ws.cell(row=row, column=1, value="LIQUIDITY RATIOS")
        ws.cell(row=row, column=1).font = Font(bold=True, size=12, color="366092")
        row += 1
        
        for name, value in ratios.get('liquidity', {}).items():
            ws.cell(row=row, column=1, value=name.replace('_', ' ').title())
            ws.cell(row=row, column=2, value=f"{value:.2f}" if value is not None else "N/A")
            row += 1
        
        row += 1
        
        # Efficiency Ratios
        ws.cell(row=row, column=1, value="EFFICIENCY RATIOS")
        ws.cell(row=row, column=1).font = Font(bold=True, size=12, color="366092")
        row += 1
        
        for name, value in ratios.get('efficiency', {}).items():
            ws.cell(row=row, column=1, value=name.replace('_', ' ').title())
            ws.cell(row=row, column=2, value=f"{value:.2f}" if value is not None else "N/A")
            row += 1
        
        row += 1
        
        # Solvency Ratios
        ws.cell(row=row, column=1, value="SOLVENCY/LEVERAGE RATIOS")
        ws.cell(row=row, column=1).font = Font(bold=True, size=12, color="366092")
        row += 1
        
        for name, value in ratios.get('solvency', {}).items():
            ws.cell(row=row, column=1, value=name.replace('_', ' ').title())
            ws.cell(row=row, column=2, value=f"{value:.2f}" if value is not None else "N/A")
            row += 1
        
        row += 1
        
        # Profitability Ratios
        ws.cell(row=row, column=1, value="PROFITABILITY RATIOS")
        ws.cell(row=row, column=1).font = Font(bold=True, size=12, color="366092")
        row += 1
        
        for name, value in ratios.get('profitability', {}).items():
            ws.cell(row=row, column=1, value=name.replace('_', ' ').title())
            ws.cell(row=row, column=2, value=f"{value:.2f}%" if value is not None and ('margin' in name or 'return' in name) else f"{value:.2f}" if value is not None else "N/A")
            row += 1
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 15


def _add_income_statement_sheet(wb: Workbook, state: EquityResearchState):
    """Add income statement sheet."""
    ws = wb.create_sheet("Income Statement")
    
    financial_statements = state.get('financial_statements', {})
    income = financial_statements.get('income_statement')
    
    if income is None or income.empty:
        ws['A1'] = "No income statement data available"
        return
    
    # Title
    ws['A1'] = "INCOME STATEMENT (₹ Crores)"
    ws['A1'].font = Font(bold=True, size=14, color="366092")
    
    # Add dataframe to sheet
    _add_dataframe_to_sheet(ws, income, start_row=3)


def _add_balance_sheet_sheet(wb: Workbook, state: EquityResearchState):
    """Add balance sheet sheet."""
    ws = wb.create_sheet("Balance Sheet")
    
    financial_statements = state.get('financial_statements', {})
    balance = financial_statements.get('balance_sheet')
    
    if balance is None or balance.empty:
        ws['A1'] = "No balance sheet data available"
        return
    
    # Title
    ws['A1'] = "BALANCE SHEET (₹ Crores)"
    ws['A1'].font = Font(bold=True, size=14, color="366092")
    
    # Add dataframe to sheet
    _add_dataframe_to_sheet(ws, balance, start_row=3)


def _add_cash_flow_sheet(wb: Workbook, state: EquityResearchState):
    """Add cash flow sheet."""
    ws = wb.create_sheet("Cash Flow")
    
    financial_statements = state.get('financial_statements', {})
    cashflow = financial_statements.get('cash_flow')
    
    if cashflow is None or cashflow.empty:
        ws['A1'] = "No cash flow data available"
        return
    
    # Title
    ws['A1'] = "CASH FLOW STATEMENT (₹ Crores)"
    ws['A1'].font = Font(bold=True, size=14, color="366092")
    
    # Add dataframe to sheet
    _add_dataframe_to_sheet(ws, cashflow, start_row=3)


def _add_stock_prices_sheet(wb: Workbook, state: EquityResearchState):
    """Add stock prices sheet."""
    ws = wb.create_sheet("Stock Prices")
    
    stock_prices = state.get('stock_prices')
    
    if stock_prices is None or stock_prices.empty:
        ws['A1'] = "No stock price data available"
        return
    
    # Title
    ws['A1'] = "HISTORICAL STOCK PRICES"
    ws['A1'].font = Font(bold=True, size=14, color="366092")
    
    # Select relevant columns
    price_df = stock_prices[['Open', 'High', 'Low', 'Close', 'Volume']].copy()
    price_df.index.name = 'Date'
    
    # Add to sheet (limit to recent data for readability)
    _add_dataframe_to_sheet(ws, price_df.tail(100), start_row=3)


def _add_dividends_sheet(wb: Workbook, state: EquityResearchState):
    """Add dividends sheet."""
    ws = wb.create_sheet("Dividends")
    
    dividends = state.get('dividends')
    
    if dividends is None or dividends.empty:
        ws['A1'] = "No dividend history available (company may not pay dividends)"
        return
    
    # Title
    ws['A1'] = "DIVIDEND HISTORY"
    ws['A1'].font = Font(bold=True, size=14, color="366092")
    
    # Add to sheet
    _add_dataframe_to_sheet(ws, dividends, start_row=3)


def _add_valuation_sheet(wb: Workbook, state: EquityResearchState):
    """Add valuation analysis sheet."""
    ws = wb.create_sheet("Valuation")
    
    # Title
    ws['A1'] = "VALUATION ANALYSIS"
    ws['A1'].font = Font(bold=True, size=14, color="366092")
    
    # Beta Analysis
    ws['A3'] = "BETA ANALYSIS"
    ws['A3'].font = Font(bold=True, size=12, color="366092")
    
    beta = state.get('beta', 0)
    correlation = state.get('correlation_with_market', 0)
    
    ws['A4'] = "Beta (vs NIFTY 50):"
    ws['B4'] = f"{beta:.3f}" if beta else "N/A"
    ws['A5'] = "Interpretation:"
    ws['B5'] = "Aggressive" if beta and beta > 1 else "Defensive" if beta else "N/A"
    ws['A6'] = "Correlation:"
    ws['B6'] = f"{correlation:.3f}" if correlation else "N/A"
    
    # CAPM
    ws['A8'] = "COST OF EQUITY (CAPM)"
    ws['A8'].font = Font(bold=True, size=12, color="366092")
    
    from config.settings import RISK_FREE_RATE, EXPECTED_MARKET_RETURN
    cost_of_equity = state.get('cost_of_equity', 0)
    
    ws['A9'] = "Risk-Free Rate:"
    ws['B9'] = f"{RISK_FREE_RATE:.2%}"
    ws['A10'] = "Expected Market Return:"
    ws['B10'] = f"{EXPECTED_MARKET_RETURN:.2%}"
    ws['A11'] = "Beta:"
    ws['B11'] = f"{beta:.3f}" if beta else "N/A"
    ws['A12'] = "Cost of Equity:"
    ws['B12'] = f"{cost_of_equity:.2%}" if cost_of_equity else "N/A"
    
    # DDM
    ws['A14'] = "DIVIDEND DISCOUNT MODEL (DDM)"
    ws['A14'].font = Font(bold=True, size=12, color="366092")
    
    ddm = state.get('ddm_valuation', {})
    
    if ddm and ddm.get('applicable'):
        ws['A15'] = "Current Dividend (D0):"
        ws['B15'] = f"₹{ddm.get('d0_current_dividend', 0):.2f}"
        ws['A16'] = "Next Dividend (D1):"
        ws['B16'] = f"₹{ddm.get('d1_next_dividend', 0):.2f}"
        ws['A17'] = "Growth Rate:"
        ws['B17'] = f"{ddm.get('growth_rate', 0):.2%}"
        ws['A18'] = "Fair Value:"
        ws['B18'] = f"₹{ddm.get('fair_value', 0):.2f}"
        
        stock_prices = state.get('stock_prices')
        if stock_prices is not None and not stock_prices.empty:
            current_price = stock_prices['Close'].iloc[-1]
            ws['A19'] = "Current Price:"
            ws['B19'] = f"₹{current_price:.2f}"
            ws['A20'] = "Upside/Downside:"
            ws['B20'] = f"{ddm.get('upside_downside', 0):.1%}"
    else:
        ws['A15'] = "DDM Not Applicable:"
        ws['B15'] = ddm.get('reason', 'Company does not pay dividends')
    
    # Style labels
    for row in range(4, 21):
        ws.cell(row=row, column=1).font = Font(bold=True)
    
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 20


def _add_news_sheet(wb: Workbook, state: EquityResearchState):
    """Add news/developments sheet."""
    ws = wb.create_sheet("News")
    
    news = state.get('news')
    
    if news is None or news.empty:
        ws['A1'] = "No news data available"
        return
    
    # Title
    ws['A1'] = "RECENT NEWS & DEVELOPMENTS"
    ws['A1'].font = Font(bold=True, size=14, color="366092")
    
    # Prepare news dataframe
    news_df = news[['published', 'title', 'source']].copy()
    news_df['published'] = news_df['published'].dt.strftime('%Y-%m-%d')
    news_df.columns = ['Date', 'Headline', 'Source']
    
    # Add to sheet (limit to most recent)
    _add_dataframe_to_sheet(ws, news_df.head(50), start_row=3)


def _add_dataframe_to_sheet(ws, df: pd.DataFrame, start_row: int = 1):
    """Add a pandas DataFrame to worksheet with formatting."""
    # Convert DataFrame to rows
    for r_idx, row in enumerate(dataframe_to_rows(df, index=True, header=True), start_row):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx)
            
            # Format value
            if pd.notna(value):
                if isinstance(value, (int, float)):
                    if abs(value) > 1e6:
                        cell.value = value / 1e7
                        cell.number_format = '#,##0.00'
                    else:
                        cell.value = value
                        cell.number_format = '#,##0.00' if isinstance(value, float) else '0'
                else:
                    cell.value = str(value)
            
            # Header row styling
            if r_idx == start_row + 1:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
    
    # Adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width


if __name__ == "__main__":
    """Test Excel workbook generation."""
    print("Testing Excel Workbook Generator...")
    
    from agents.state import create_initial_state
    from agents.nodes import collect_data_node, analyze_node
    
    test_ticker = "RELIANCE"
    
    try:
        print(f"\n🧪 Testing with {test_ticker}...")
        
        # Collect data and analyze
        print("📊 Collecting data...")
        state = create_initial_state(test_ticker, "Reliance Industries")
        data_updates = collect_data_node(state)
        state.update(data_updates)
        
        print("📈 Analyzing...")
        analysis_updates = analyze_node(state)
        state.update(analysis_updates)
        
        # Generate Excel workbook
        print("\n📊 Generating Excel workbook...")
        filepath = generate_excel_workbook(state)
        
        print(f"\n✅ Test completed!")
        print(f"   Workbook saved: {filepath}")
        
    except Exception as e:
        print(f"\n❌ Test failed: {e}")
        import traceback
        traceback.print_exc()

